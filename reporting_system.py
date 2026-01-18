# reporting_system.py
"""
POS Reporting System - Clean, Non-Redundant Sales Reports
Handles both daily cashier reports and monthly admin summaries
"""

import sqlite3
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
import os
import sys
try:
    from openpyxl.cell.cell import MergedCell
except Exception:
    MergedCell = None
from openpyxl.utils import get_column_letter

# Central DB name resolution (keeps compatibility with sales_utils DB selection)
DB_NAME = os.environ.get('BAR_SALES_DB') or os.path.join(os.getcwd(), 'bar_sales.db')
if 'unittest' in sys.modules and 'BAR_SALES_DB' not in os.environ:
    DB_NAME = 'test_bar_sales.db'


def get_db_connection():
    """Get database connection with safe settings"""
    conn = sqlite3.connect(DB_NAME, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn


# ========================================
# PART 1 & 2: CASHIER DAILY SALES REPORT
# ========================================

def get_daily_sales_summary(selected_date: str) -> Dict:
    """
    Get sales summary for a single specific date only.
    Args:
        selected_date: Date in 'YYYY-MM-DD' format
    Returns:
        Dictionary with total_sales, transaction_count, payment_breakdown, sales_list
    """
    conn = get_db_connection()
    cur = conn.cursor()

    # Define exact date range (00:00:00 to 23:59:59)
    start_ts = f"{selected_date} 00:00:00"
    end_ts = f"{selected_date} 23:59:59"

    try:
        # Total sales amount (excluding voided)
        cur.execute("""
            SELECT COALESCE(SUM(total), 0) as total_amount
            FROM sales 
            WHERE status != 'VOIDED' 
            AND timestamp BETWEEN ? AND ?
        """, (start_ts, end_ts))
        total_sales = float(cur.fetchone()['total_amount'])

        # Transaction count
        cur.execute("""
            SELECT COUNT(*) as tx_count
            FROM sales 
            WHERE status != 'VOIDED' 
            AND timestamp BETWEEN ? AND ?
        """, (start_ts, end_ts))
        transaction_count = cur.fetchone()['tx_count']

        # Payment method breakdown
        cur.execute("""
            SELECT 
                payment_method,
                COUNT(*) as count,
                COALESCE(SUM(total), 0) as amount
            FROM sales 
            WHERE status != 'VOIDED' 
            AND timestamp BETWEEN ? AND ?
            GROUP BY payment_method
        """, (start_ts, end_ts))

        payment_breakdown = {}
        for row in cur.fetchall():
            method = row['payment_method'] or 'Cash'
            payment_breakdown[method] = {
                'count': row['count'],
                'amount': float(row['amount'])
            }

        # Sales list with details
        cur.execute("""
            SELECT 
                transaction_id,
                timestamp,
                total,
                payment_method,
                cashier
            FROM sales 
            WHERE status != 'VOIDED' 
            AND timestamp BETWEEN ? AND ?
            ORDER BY timestamp ASC
        """, (start_ts, end_ts))

        sales_list = []
        for row in cur.fetchall():
            sales_list.append({
                'time': row['timestamp'].split(' ')[1] if ' ' in row['timestamp'] else row['timestamp'],
                'receipt_number': row['transaction_id'],
                'amount': float(row['total']),
                'payment_method': row['payment_method'] or 'Cash',
                'cashier': row['cashier']
            })

        return {
            'date': selected_date,
            'total_sales': total_sales,
            'transaction_count': transaction_count,
            'payment_breakdown': payment_breakdown,
            'sales_list': sales_list
        }

    finally:
        conn.close()


def get_daily_items_sold(selected_date: str) -> List[Tuple]:
    """
    Get items sold for a specific date only.
    Returns: List of (item_name, quantity, revenue)
    """
    conn = get_db_connection()
    cur = conn.cursor()

    start_ts = f"{selected_date} 00:00:00"
    end_ts = f"{selected_date} 23:59:59"

    try:
        cur.execute("""
            SELECT 
                si.item,
                COALESCE(SUM(si.quantity), 0) as qty_sold,
                COALESCE(SUM(si.subtotal), 0) as revenue
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'VOIDED'
            AND s.timestamp BETWEEN ? AND ?
            GROUP BY si.item
            ORDER BY revenue DESC
        """, (start_ts, end_ts))

        return [(row['item'], int(row['qty_sold']), float(row['revenue']))
                for row in cur.fetchall()]
    finally:
        conn.close()


# ========================================
# PART 3: MONTHLY SALES SUMMARY (ADMIN)
# ========================================

def get_monthly_sales_summary(year: int, month: int) -> Dict:
    """
    Get comprehensive monthly sales summary.
    Args:
        year: Year (e.g., 2025)
        month: Month (1-12)
    Returns:
        Dictionary with monthly overview, payment summary, daily breakdown, top items
    """
    from calendar import monthrange

    # Get first and last day of the month
    first_day = f"{year}-{month:02d}-01"
    last_day_num = monthrange(year, month)[1]
    last_day = f"{year}-{month:02d}-{last_day_num:02d}"

    start_ts = f"{first_day} 00:00:00"
    end_ts = f"{last_day} 23:59:59"

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # ===== SECTION A: Monthly Overview =====
        cur.execute("""
            SELECT 
                COALESCE(SUM(total), 0) as total_revenue,
                COUNT(*) as total_transactions
            FROM sales 
            WHERE status != 'VOIDED' 
            AND timestamp BETWEEN ? AND ?
        """, (start_ts, end_ts))

        row = cur.fetchone()
        total_revenue = float(row['total_revenue'])
        total_transactions = row['total_transactions']

        # Get daily breakdown for calculations
        cur.execute("""
            SELECT 
                DATE(timestamp) as sale_date,
                COUNT(*) as tx_count,
                COALESCE(SUM(total), 0) as daily_total
            FROM sales 
            WHERE status != 'VOIDED' 
            AND timestamp BETWEEN ? AND ?
            GROUP BY DATE(timestamp)
            ORDER BY sale_date ASC
        """, (start_ts, end_ts))

        daily_breakdown = []
        best_day = {'date': 'N/A', 'amount': 0}
        worst_day = {'date': 'N/A', 'amount': float('inf')}

        for row in cur.fetchall():
            date = row['sale_date']
            tx_count = row['tx_count']
            daily_total = float(row['daily_total'])

            daily_breakdown.append({
                'date': date,
                'transaction_count': tx_count,
                'daily_total': daily_total
            })

            if daily_total > best_day['amount']:
                best_day = {'date': date, 'amount': daily_total}

            if daily_total < worst_day['amount'] and daily_total > 0:
                worst_day = {'date': date, 'amount': daily_total}

        # Calculate average
        num_days_with_sales = len(daily_breakdown)
        avg_daily_sales = total_revenue / num_days_with_sales if num_days_with_sales > 0 else 0

        # ===== SECTION B: Payment Summary =====
        cur.execute("""
            SELECT 
                payment_method,
                COALESCE(SUM(total), 0) as amount
            FROM sales 
            WHERE status != 'VOIDED' 
            AND timestamp BETWEEN ? AND ?
            GROUP BY payment_method
        """, (start_ts, end_ts))

        payment_summary = {}
        for row in cur.fetchall():
            method = row['payment_method'] or 'Cash'
            amount = float(row['amount'])
            percentage = (amount / total_revenue * 100) if total_revenue > 0 else 0
            payment_summary[method] = {
                'amount': amount,
                'percentage': percentage
            }

        # ===== SECTION D: Top Selling Items =====
        cur.execute("""
            SELECT 
                si.item,
                COALESCE(SUM(si.quantity), 0) as qty_sold,
                COALESCE(SUM(si.subtotal), 0) as revenue
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE s.status != 'VOIDED'
            AND s.timestamp BETWEEN ? AND ?
            GROUP BY si.item
            ORDER BY revenue DESC
            LIMIT 20
        """, (start_ts, end_ts))

        top_items = []
        for row in cur.fetchall():
            top_items.append({
                'item_name': row['item'],
                'quantity_sold': int(row['qty_sold']),
                'revenue': float(row['revenue'])
            })

        return {
            'month': f"{year}-{month:02d}",
            'month_name': datetime(year, month, 1).strftime('%B %Y'),
            'overview': {
                'total_revenue': total_revenue,
                'total_transactions': total_transactions,
                'avg_daily_sales': avg_daily_sales,
                'best_day': best_day,
                'worst_day': worst_day if worst_day['amount'] != float('inf') else {'date': 'N/A', 'amount': 0}
            },
            'payment_summary': payment_summary,
            'daily_breakdown': daily_breakdown,
            'top_items': top_items
        }

    finally:
        conn.close()


# ========================================
# PART 4: EXPORT FUNCTIONS
# ========================================

def export_daily_sales_to_excel(selected_date: str, include_expenses: bool = True):
    """
    Export daily cashier report to Excel
    File name: Daily_Sales_YYYY_MM_DD.xlsx
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Load business settings for header
    try:
        import business_settings
        receipt_settings = business_settings.get_receipt_settings()
        business_name = receipt_settings.get('business_name', 'Gorgeous Brides Boutique') or 'Gorgeous Brides Boutique'
        currency_symbol = receipt_settings.get('currency_symbol', 'ZMW') or 'ZMW'
        phone_primary = receipt_settings.get('phone_primary', '')
        email = receipt_settings.get('email', '')
        address = receipt_settings.get('address', '')
        tpin = receipt_settings.get('tpin', '')
    except Exception:
        business_name = 'Gorgeous Brides Boutique'
        currency_symbol = 'ZMW'
        phone_primary = ''
        email = ''
        address = ''
        tpin = ''

    # Get sales data
    summary = get_daily_sales_summary(selected_date)
    items = get_daily_items_sold(selected_date)

    # Get expenses if requested
    expenses = []
    total_expenses = 0.0
    if include_expenses:
        try:
            from expenses_system import get_expenses_by_date_range
            expenses_raw = get_expenses_by_date_range(selected_date, selected_date)
            for exp in expenses_raw:
                if len(exp) >= 8 and isinstance(exp[0], int):
                    _, date, category, description, amount_val, cashier_name, created_at, notes = exp[:8]
                else:
                    try:
                        date, category, description, amount_val, cashier_name, created_at, notes = tuple(exp[:7])
                    except:
                        continue

                try:
                    amount = float(amount_val) if amount_val is not None else 0.0
                except:
                    amount = 0.0

                expenses.append({
                    'date': date,
                    'category': category or '',
                    'description': description or '',
                    'amount': amount,
                    'cashier': cashier_name or '',
                    'notes': notes or ''
                })
                total_expenses += amount
        except:
            pass

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Daily Sales {selected_date}"

    # Styling
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    current_row = 1

    # Business Name Header
    ws.merge_cells(f'A{current_row}:E{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = business_name.upper()
    cell.font = Font(name='Calibri', size=16, bold=True, color='2E5C8A')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    # Contact info row (optional)
    contact_parts = []
    if phone_primary:
        contact_parts.append(f"Tel: {phone_primary}")
    if email:
        contact_parts.append(f"Email: {email}")
    if tpin:
        contact_parts.append(f"TPIN: {tpin}")

    if contact_parts:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = " | ".join(contact_parts)
        cell.font = Font(name='Calibri', size=9, color='7F8C8D')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    current_row += 1  # Add spacing

    # Title
    ws.merge_cells(f'A{current_row}:E{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f'DAILY SALES REPORT - {selected_date}'
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2

    # Sales Summary Section
    ws[f'A{current_row}'] = 'SALES SUMMARY'
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    ws[f'A{current_row}'] = 'Total Sales:'
    ws[f'B{current_row}'] = summary['total_sales']
    ws[f'B{current_row}'].number_format = f'"{currency_symbol} "#,##0.00'
    current_row += 1

    ws[f'A{current_row}'] = 'Number of Transactions:'
    ws[f'B{current_row}'] = summary['transaction_count']
    current_row += 2

    # Payment Method Breakdown
    ws[f'A{current_row}'] = 'Payment Method Breakdown:'
    ws[f'A{current_row}'].font = Font(bold=True)
    current_row += 1

    for method, data in summary['payment_breakdown'].items():
        ws[f'A{current_row}'] = f'  {method}:'
        ws[f'B{current_row}'] = data['amount']
        ws[f'B{current_row}'].number_format = f'"{currency_symbol} "#,##0.00'
        ws[f'C{current_row}'] = f"({data['count']} transactions)"
        current_row += 1

    current_row += 1

    # Sales List
    ws[f'A{current_row}'] = 'SALES TRANSACTIONS'
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    headers = ['Time', 'Receipt Number', 'Amount', 'Payment Method', 'Cashier']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    current_row += 1

    for sale in summary['sales_list']:
        ws.cell(row=current_row, column=1, value=sale['time'])
        ws.cell(row=current_row, column=2, value=sale['receipt_number'])
        cell = ws.cell(row=current_row, column=3, value=sale['amount'])
        cell.number_format = f'"{currency_symbol} "#,##0.00'
        ws.cell(row=current_row, column=4, value=sale['payment_method'])
        ws.cell(row=current_row, column=5, value=sale['cashier'])
        current_row += 1

    current_row += 1

    # Items Sold
    ws[f'A{current_row}'] = 'ITEMS SOLD'
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    headers = ['Item Name', 'Quantity Sold', 'Revenue']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    current_row += 1

    for item_name, qty, revenue in items:
        ws.cell(row=current_row, column=1, value=item_name)
        ws.cell(row=current_row, column=2, value=qty)
        cell = ws.cell(row=current_row, column=3, value=revenue)
        cell.number_format = f'"{currency_symbol} "#,##0.00'
        current_row += 1

    current_row += 1

    # Expenses Section (if included)
    if include_expenses and expenses:
        ws[f'A{current_row}'] = 'DAILY EXPENSES'
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1

        headers = ['Date', 'Category', 'Description', 'Amount', 'Cashier', 'Notes']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        current_row += 1

        for exp in expenses:
            ws.cell(row=current_row, column=1, value=exp['date'])
            ws.cell(row=current_row, column=2, value=exp['category'])
            ws.cell(row=current_row, column=3, value=exp['description'])
            cell = ws.cell(row=current_row, column=4, value=exp['amount'])
            cell.number_format = f'"{currency_symbol} "#,##0.00'
            ws.cell(row=current_row, column=5, value=exp['cashier'])
            ws.cell(row=current_row, column=6, value=exp['notes'])
            current_row += 1

        current_row += 1
        ws[f'A{current_row}'] = 'TOTAL EXPENSES:'
        ws[f'A{current_row}'].font = Font(bold=True)
        cell = ws[f'B{current_row}']
        cell.value = total_expenses
        cell.number_format = f'"{currency_symbol} "#,##0.00'
        cell.font = Font(bold=True)
        current_row += 2

    # Final Summary
    ws[f'A{current_row}'] = 'NET TOTAL:'
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    net_total = summary['total_sales'] - total_expenses
    cell = ws[f'B{current_row}']
    cell.value = net_total
    cell.number_format = f'"{currency_symbol} "#,##0.00'
    cell.font = Font(bold=True, size=12, color='27AE60' if net_total >= 0 else 'C0392B')

    # Auto-size columns safely: iterate by column index and compute max length of visible cell values
    try:
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Skip merged cell proxies which may not have .value
                    val = cell.value
                    if val is None:
                        continue
                    s = str(val)
                    if len(s) > max_length:
                        max_length = len(s)
                except Exception:
                    continue
            # Apply width with a small padding
            if max_length > 0:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max_length + 2
    except Exception:
        # If anything goes wrong, ignore auto-sizing to avoid breaking export
        pass

    # Save file
    os.makedirs('exports', exist_ok=True)
    filename = f"Daily_Sales_{selected_date.replace('-', '_')}.xlsx"
    filepath = os.path.join('exports', filename)
    wb.save(filepath)
    return filepath, summary['total_sales'], total_expenses, net_total


def export_monthly_sales_to_excel(year: int, month: int):
    """
    Export monthly sales summary to Excel
    File name: Monthly_Sales_Report_YYYY_MM.xlsx
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Load business settings for header
    try:
        import business_settings
        receipt_settings = business_settings.get_receipt_settings()
        business_name = receipt_settings.get('business_name', 'Gorgeous Brides Boutique') or 'Gorgeous Brides Boutique'
        currency_symbol = receipt_settings.get('currency_symbol', 'ZMW') or 'ZMW'
        phone_primary = receipt_settings.get('phone_primary', '')
        email = receipt_settings.get('email', '')
        tpin = receipt_settings.get('tpin', '')
    except Exception:
        business_name = 'Gorgeous Brides Boutique'
        currency_symbol = 'ZMW'
        phone_primary = ''
        email = ''
        tpin = ''

    summary = get_monthly_sales_summary(year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Monthly Sales {year}-{month:02d}"

    # Styling
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    title_fill = PatternFill(start_color='2E5C8A', end_color='2E5C8A', fill_type='solid')
    section_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

    current_row = 1

    # Business Name Header
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = business_name.upper()
    cell.font = Font(name='Calibri', size=18, bold=True, color='2E5C8A')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1

    # Contact info row (optional)
    contact_parts = []
    if phone_primary:
        contact_parts.append(f"Tel: {phone_primary}")
    if email:
        contact_parts.append(f"Email: {email}")
    if tpin:
        contact_parts.append(f"TPIN: {tpin}")

    if contact_parts:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = " | ".join(contact_parts)
        cell.font = Font(name='Calibri', size=9, color='7F8C8D')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

    current_row += 1  # Add spacing

    # Main Title
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f'MONTHLY SALES REPORT - {summary["month_name"]}'
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[current_row].height = 25
    current_row += 2

    # ===== SECTION A: Monthly Overview =====
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'SECTION A — MONTHLY OVERVIEW'
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1

    overview = summary['overview']
    ws[f'A{current_row}'] = 'Total Monthly Revenue:'
    cell = ws[f'B{current_row}']
    cell.value = overview['total_revenue']
    cell.number_format = f'"{currency_symbol} "#,##0.00'
    current_row += 1

    ws[f'A{current_row}'] = 'Total Transactions:'
    ws[f'B{current_row}'] = overview['total_transactions']
    current_row += 1

    ws[f'A{current_row}'] = 'Average Daily Sales:'
    cell = ws[f'B{current_row}']
    cell.value = overview['avg_daily_sales']
    cell.number_format = f'"{currency_symbol} "#,##0.00'
    current_row += 1

    ws[f'A{current_row}'] = 'Best Performing Day:'
    ws[f'B{current_row}'] = overview['best_day']['date']
    cell = ws[f'C{current_row}']
    cell.value = overview['best_day']['amount']
    cell.number_format = f'"{currency_symbol} "#,##0.00'
    current_row += 1

    ws[f'A{current_row}'] = 'Lowest Performing Day:'
    ws[f'B{current_row}'] = overview['worst_day']['date']
    cell = ws[f'C{current_row}']
    cell.value = overview['worst_day']['amount']
    cell.number_format = f'"{currency_symbol} "#,##0.00'
    current_row += 2

    # ===== SECTION B: Payment Summary =====
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'SECTION B — PAYMENT METHOD SUMMARY'
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1

    headers = ['Payment Method', 'Total Amount', 'Percentage']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    for method, data in summary['payment_summary'].items():
        ws.cell(row=current_row, column=1, value=method)
        cell = ws.cell(row=current_row, column=2, value=data['amount'])
        cell.number_format = f'"{currency_symbol} "#,##0.00'
        cell = ws.cell(row=current_row, column=3, value=data['percentage'])
        cell.number_format = '0.00"%"'
        current_row += 1

    current_row += 1

    # ===== SECTION C: Daily Breakdown =====
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'SECTION C — DAILY BREAKDOWN'
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1

    headers = ['Date', 'Number of Sales', 'Daily Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    for day in summary['daily_breakdown']:
        ws.cell(row=current_row, column=1, value=day['date'])
        ws.cell(row=current_row, column=2, value=day['transaction_count'])
        cell = ws.cell(row=current_row, column=3, value=day['daily_total'])
        cell.number_format = f'"{currency_symbol} "#,##0.00'
        current_row += 1

    current_row += 1

    # ===== SECTION D: Top Selling Items =====
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = 'SECTION D — TOP SELLING ITEMS'
    cell.font = section_font
    cell.fill = section_fill
    current_row += 1

    headers = ['Item Name', 'Quantity Sold', 'Revenue']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    for item in summary['top_items']:
        ws.cell(row=current_row, column=1, value=item['item_name'])
        ws.cell(row=current_row, column=2, value=item['quantity_sold'])
        cell = ws.cell(row=current_row, column=3, value=item['revenue'])
        cell.number_format = f'"{currency_symbol} "#,##0.00'
        current_row += 1

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save
    os.makedirs('exports', exist_ok=True)
    filename = f"Monthly_Sales_Report_{year}_{month:02d}.xlsx"
    filepath = os.path.join('exports', filename)
    wb.save(filepath)
    # Auto-size columns - robust handling for merged cells / different openpyxl versions
    return filepath, overview['total_revenue']
